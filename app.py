"""Assortment & Allocation Automation App — Flask routes & BigQuery logic."""

import logging
import os
import io
import re
import math
import datetime

import pandas as pd
from flask import Flask, render_template, jsonify, request, send_file
from google.cloud import bigquery
import google.auth
import google.auth.transport.requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import (
    PROJECT_ID, TEMP_DATASET, EVENTS_SKU_LIST, VENDOR_STRATEGY,
    DFC_COST_MODEL_SUBMISSION, FINAL_ALLOCATIONS, SCHN_SKU_ATTR,
    TEMPLATE_COLUMNS_DOMESTIC, TEMPLATE_COLUMNS_IMPORT,
    CONTAINER_DIVISOR, STRATEGY_KEYS, MAX_UPLOAD_MB,
    ALLOWED_DFCS, DC_NAMES, CATALOG_RUN_ANALYTICS,
)
from validators import validate_upload, _determine_thd_key
from assortment_engine import determine_assortment_ids, start_multi_dc, fetch_multi_dc_results
from allocation_engine import run_allocation, fetch_results, fetch_summary, validate_results, fetch_available_dc_counts, fetch_factory_summary
from event_history import fetch_prior_year_strategy, fetch_known_event_names

app = Flask(__name__)
app.config["JSON_SORT_KEYS"] = False
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
app.config["TEMPLATES_AUTO_RELOAD"] = True

@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

CURRENT_USER = os.environ.get("USERNAME", os.environ.get("USER", "unknown"))

_bq_client = None
_bq_creds = None
_upload_cache = {}


def _load_and_refresh_creds():
    """Load ADC from disk and refresh the access token."""
    creds, _ = google.auth.default()
    creds.refresh(google.auth.transport.requests.Request())
    return creds


def bq():
    """Return a BigQuery client, reloading credentials from disk each call."""
    global _bq_client, _bq_creds, CURRENT_USER
    try:
        creds = _load_and_refresh_creds()
    except Exception as exc:
        raise RuntimeError(
            "Reauthentication is needed. Please run "
            "`gcloud auth application-default login` to reauthenticate."
        ) from exc

    # Reuse client if token unchanged
    if _bq_creds is not None and getattr(_bq_creds, 'token', None) == creds.token:
        return _bq_client

    _bq_creds = creds
    _bq_client = bigquery.Client(project=PROJECT_ID, credentials=_bq_creds)
    if hasattr(creds, "service_account_email"):
        CURRENT_USER = creds.service_account_email
    logger.info(f"Authenticated as: {CURRENT_USER}")
    return _bq_client


# ── Routes ──────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/user_info")
def api_user_info():
    email = f"{CURRENT_USER}@homedepot.com"
    try:
        q = """SELECT email_addr_txt AS email FROM `pr-edw-views-thd.ASSOC.ASSOC`
               WHERE actv_flg = 'Y' AND email_addr_txt IS NOT NULL AND user_id = @ldap LIMIT 1"""
        jc = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("ldap", "STRING", CURRENT_USER),
        ])
        rows = list(bq().query(q, job_config=jc).result())
        if rows:
            email = rows[0].email
    except Exception:
        pass
    return jsonify({"ldap_user_id": CURRENT_USER, "email": email})


@app.route("/api/config")
def api_config():
    return jsonify({
        "strategies": STRATEGY_KEYS,
        "dc_count_options": DC_COUNT_OPTIONS,
        "allowed_dfcs": ALLOWED_DFCS,
        "dc_names": {str(k): v for k, v in DC_NAMES.items()},
    })


# ── Section 1: Template Download ────────────────────────────────────

def _build_template(columns, filename, event_name="", event_year=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "SKU Upload Template"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    req_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for i, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=col["name"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(len(col["name"]) + 4, 14)

    info_font = Font(italic=True, size=8, color="666666")
    for i, col in enumerate(columns, 1):
        label = col["type"]
        if col.get("note"):
            label += f" ({col['note']})"
        elif col["required"]:
            label += " (Required)"
        cell = ws.cell(row=2, column=i, value=label)
        cell.font = info_font
        cell.border = thin_border
        if col["required"] or col.get("note"):
            cell.fill = req_fill

    if event_name:
        ws.cell(row=2, column=1, value=event_name)
    if event_year and len(columns) > 1 and columns[1]["name"] == "EVENT_YEAR":
        ws.cell(row=2, column=2, value=int(event_year) if str(event_year).isdigit() else event_year)

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/download_template")
def download_template():
    includes_imports = request.args.get("imports", "false").lower() == "true"
    event_name = request.args.get("event_name", "").strip()
    event_year = request.args.get("event_year", "").strip()
    if includes_imports:
        return _build_template(TEMPLATE_COLUMNS_IMPORT, "sku_upload_template_import.xlsx", event_name, event_year)
    return _build_template(TEMPLATE_COLUMNS_DOMESTIC, "sku_upload_template_domestic.xlsx", event_name, event_year)


@app.route("/api/upload_status")
def api_upload_status():
    """The authoritative import/domestic flag — this session's own upload if there
    was one, else what's already persisted in EVENTS_SKU_LIST for this event (so
    reusing an event that was uploaded in an earlier session, without re-uploading,
    still resolves correctly). Returns null when neither source has an answer, so
    the client knows to keep whatever it already has rather than being told
    "domestic" by default."""
    event_name = request.args.get("event_name", "")
    return jsonify({"includes_imports": _resolve_is_import(event_name)})


@app.route("/api/resolve_sku_grp")
def api_resolve_sku_grp():
    """The real SKU_GRP for a Run ID + Event Name, straight from the catalog run —
    never guessed client-side. Returns null if the run doesn't have a matching key
    yet (e.g. the catalog tool hasn't been run for this event)."""
    run_id = request.args.get("run_id", "")
    event_name = request.args.get("event_name", "")
    return jsonify({"sku_grp": _resolve_sku_grp(run_id, event_name)})


@app.route("/api/known_event_names")
def api_known_event_names():
    """Canonical event names already recorded in the historical allocation
    table — backs Step 1's dropdown so a new entry stays consistent with
    existing ones instead of fragmenting into a near-duplicate free-text
    variant."""
    return jsonify({"event_names": fetch_known_event_names(bq())})


@app.route("/api/prior_year_strategy")
def api_prior_year_strategy():
    """Most recent recorded strategy snapshot at or before the year prior to
    the one being planned, from EVENTS_SKU_DFC_ALLOCATIONS."""
    event_name = request.args.get("event_name", "")
    event_year = request.args.get("event_year", "")
    is_import = request.args.get("is_import", "")
    if not event_name or not event_year:
        return jsonify({"error": "event_name and event_year are required"}), 400
    try:
        year_int = int(event_year)
    except ValueError:
        return jsonify({"error": "event_year must be a number"}), 400
    result = fetch_prior_year_strategy(bq(), event_name, year_int - 1, is_import)
    return jsonify(result)


# ── Section 2: File Upload & Validation ─────────────────────────────

@app.route("/api/upload", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400

    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in ("xlsx", "csv"):
        return jsonify({"error": "Only .xlsx and .csv files are accepted"}), 400

    try:
        raw = io.BytesIO(f.read())
        if ext == "csv":
            df = pd.read_csv(raw, dtype=str)
        else:
            df = pd.read_excel(raw, engine="openpyxl", dtype=str)
    except Exception as e:
        logger.exception("File parse error")
        return jsonify({"error": f"Failed to parse file: {e}"}), 400

    df.columns = [c.strip().upper() for c in df.columns]

    # Strip template type-hint row if present (row 2 of downloaded template)
    if len(df) > 0:
        first = df.iloc[0].astype(str).str.strip().str.upper()
        type_keywords = {"STRING", "INT64", "FLOAT64", "REQUIRED", "OPTIONAL"}
        matches = first.isin(type_keywords).sum()
        if matches >= len(df.columns) // 2:
            df = df.iloc[1:].reset_index(drop=True)

    includes_imports = request.form.get("includes_imports", "false").lower() == "true"
    result = validate_upload(df, includes_imports=includes_imports)

    if result["passed"]:
        _upload_cache["df"] = df
        _upload_cache["event_name"] = result["summary"].get("event_name", "")
        _upload_cache["includes_imports"] = includes_imports
        if "EVENT_YEAR" in df.columns and len(df) > 0:
            _upload_cache["event_year"] = int(df["EVENT_YEAR"].iloc[0])
            result["summary"]["event_year"] = _upload_cache["event_year"]
        # Detect wave count from columns with non-null data
        wave_cols = [c for c in ["WAVE_1", "WAVE_2", "WAVE_3", "WAVE_4"] if c in df.columns and df[c].notna().any()]
        _upload_cache["wave_count"] = len(wave_cols)
        result["summary"]["wave_count"] = len(wave_cols)

        # Pre-populate factory cubes from existing BQ data
        if includes_imports and result["summary"].get("factory_distribution"):
            event_name = result["summary"].get("event_name", "")
            event_year = _upload_cache.get("event_year")
            if event_name and event_year is not None:
                try:
                    fc_query = f"""
                        SELECT FACTORY_ID, MAX(FACTORY_CUBE) AS FACTORY_CUBE
                        FROM {EVENTS_SKU_LIST}
                        WHERE UPPER(EVENT_NAME) = @ev AND EVENT_YEAR = @yr
                          AND FACTORY_ID IS NOT NULL
                        GROUP BY FACTORY_ID
                    """
                    fc_cfg = bigquery.QueryJobConfig(query_parameters=[
                        bigquery.ScalarQueryParameter("ev", "STRING", event_name.upper()),
                        bigquery.ScalarQueryParameter("yr", "INT64", event_year),
                    ])
                    fc_map = {int(r.FACTORY_ID): float(r.FACTORY_CUBE or 0)
                              for r in bq().query(fc_query, job_config=fc_cfg).result()}
                    if fc_map:
                        for f in result["summary"]["factory_distribution"]:
                            f["factory_cube"] = fc_map.get(f["factory_id"], 0)
                        _upload_cache["factory_cubes"] = fc_map
                except Exception:
                    logger.exception("Pre-populate factory cubes failed")

    # Detect mismatch: user chose domestic but file has FACTORY_ID
    if not includes_imports and "FACTORY_ID" in df.columns:
        result["import_mismatch"] = True
        result["import_mismatch_msg"] = (
            "This file contains a FACTORY_ID column, which indicates an import file. "
            "Please go back to Step 1 and select 'Import Only' with your container size."
        )

    return jsonify(result)


# ── Section 2b: BigQuery Validation ─────────────────────────────────

def _sanitize_table_name(s):
    return re.sub(r'[^A-Za-z0-9_]', '_', s.strip().upper())


@app.route("/api/validate_bq", methods=["POST"])
def validate_bq():
    df = _upload_cache.get("df")
    if df is None:
        return jsonify({"error": "No validated data. Please upload first."}), 400

    event_name = _upload_cache.get("event_name", "")
    event_year = _upload_cache.get("event_year")
    if not event_name or event_year is None:
        return jsonify({"error": "EVENT_NAME and EVENT_YEAR are required."}), 400

    checks = []

    # 1. Check EVENT_NAME + EVENT_YEAR not already in EVENTS_SKU_LIST
    try:
        dup_query = f"""
            SELECT COUNT(*) AS cnt
            FROM {EVENTS_SKU_LIST}
            WHERE UPPER(EVENT_NAME) = @event_name AND EVENT_YEAR = @event_year
        """
        job_config = bigquery.QueryJobConfig(
            query_parameters=[
                bigquery.ScalarQueryParameter("event_name", "STRING", event_name.upper()),
                bigquery.ScalarQueryParameter("event_year", "INT64", event_year),
            ]
        )
        cnt = next(iter(bq().query(dup_query, job_config=job_config).result())).cnt
        event_exists = cnt > 0
        checks.append({
            "name": "Event not in EVENTS_SKU_LIST",
            "passed": not event_exists,
            "detail": f"{event_name} {event_year} already has {cnt:,} rows"
                      if event_exists
                      else f"{event_name} {event_year} is new \u2014 ready to insert",
        })
        if event_exists:
            return jsonify({"passed": False, "checks": checks})
    except Exception as e:
        logger.exception("Event duplicate check error")
        return jsonify({"error": f"BigQuery error checking event: {e}"}), 500

    # 2. Check SKU creation dates (must be >= 365 days old)
    try:
        all_skus = set()
        if "THD_SKU_NBR" in df.columns:
            all_skus.update(int(x) for x in df["THD_SKU_NBR"].dropna())
        if "SISTER_SKU_NBR" in df.columns:
            all_skus.update(int(x) for x in df["SISTER_SKU_NBR"].dropna())
        all_skus.discard(0)

        invalid_rows = []

        if all_skus:
            sku_list = sorted(all_skus)
            age_query = f"""
                SELECT CAST(SKU_NBR AS INT64) AS SKU_NBR, SKU_CRT_DT
                FROM {SCHN_SKU_ATTR}
                WHERE SKU_NBR IN UNNEST(@sku_list)
                  AND LATEST_SKU_CRT_DT_FLG = TRUE
            """
            jc = bigquery.QueryJobConfig(
                query_parameters=[
                    bigquery.ArrayQueryParameter("sku_list", "INT64", sku_list),
                ]
            )
            age_df = bq().query(age_query, job_config=jc).to_dataframe()
            sku_dates = {}
            for _, r in age_df.iterrows():
                dt = r["SKU_CRT_DT"]
                if isinstance(dt, pd.Timestamp):
                    dt = dt.date()
                sku_dates[int(r["SKU_NBR"])] = dt

            cutoff = datetime.date.today() - datetime.timedelta(days=365)

            for idx, row in df.iterrows():
                thd = int(row["THD_SKU_NBR"]) if pd.notna(row.get("THD_SKU_NBR")) else None
                sis = int(row["SISTER_SKU_NBR"]) if pd.notna(row.get("SISTER_SKU_NBR")) else None
                row_dict = {}
                for c in df.columns:
                    v = row[c]
                    if pd.notna(v):
                        row_dict[c] = int(v) if isinstance(v, float) and v == int(v) else v

                thd_ok = thd and thd in sku_dates and sku_dates[thd] <= cutoff
                sis_ok = sis and sis in sku_dates and sku_dates[sis] <= cutoff

                # Row is valid if at least one SKU has >= 365 days
                if not thd_ok and not sis_ok:
                    reason_parts = []
                    if thd and thd in sku_dates:
                        reason_parts.append(f"THD_SKU_NBR {thd} created {sku_dates[thd]}")
                    elif thd:
                        reason_parts.append(f"THD_SKU_NBR {thd} not found in SCHN_SKU_ATTR")
                    if sis and sis in sku_dates:
                        reason_parts.append(f"SISTER_SKU_NBR {sis} created {sku_dates[sis]}")
                    elif sis:
                        reason_parts.append(f"SISTER_SKU_NBR {sis} not found in SCHN_SKU_ATTR")
                    invalid_rows.append({
                        **row_dict,
                        "INVALID_COLUMN": "THD_SKU_NBR / SISTER_SKU_NBR",
                        "INVALID_SKU": thd or sis,
                        "REASON": "Neither SKU has >= 365 days: " + "; ".join(reason_parts),
                    })

        checks.append({
            "name": "SKU age \u2265 365 days",
            "passed": len(invalid_rows) == 0,
            "detail": f"{len(invalid_rows)} row(s) where neither THD nor SISTER SKU has \u2265 365 days"
                      if invalid_rows
                      else "All rows have at least one SKU with \u2265 365 days history",
        })

        if invalid_rows:
            _upload_cache["invalid_skus"] = pd.DataFrame(invalid_rows)
            return jsonify({
                "passed": False,
                "checks": checks,
                "invalid_sku_count": len(invalid_rows),
                "has_download": True,
            })
    except Exception as e:
        logger.exception("SKU age check error")
        return jsonify({"error": f"BigQuery error checking SKU age: {e}"}), 500

    # 3. All checks passed — create temp validation table and compute factory cube
    try:
        safe_event = _sanitize_table_name(event_name)
        table_id = f"{PROJECT_ID}.{TEMP_DATASET}.VALIDATION_{safe_event}_{event_year}"
        jc = bigquery.LoadJobConfig(
            write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
        )
        bq().load_table_from_dataframe(df, table_id, job_config=jc).result()
        _upload_cache["validation_table"] = table_id

        # Compute factory cube via SCHN_SKU_ATTR (only for imports with FACTORY_ID)
        includes_imports = _upload_cache.get("includes_imports", False)
        if includes_imports and "FACTORY_ID" in df.columns:
            cube_query = f"""
                WITH src AS (
                  SELECT SAFE_CAST(THD_SKU_NBR AS INT64) AS THD_SKU_NBR,
                         SAFE_CAST(FACTORY_ID AS INT64) AS FACTORY_ID,
                         SAFE_CAST(BUY_UNITS AS INT64) AS BUY_UNITS,
                         SAFE_CAST(BP AS INT64) AS BP
                  FROM `{table_id}`
                ),
                attr AS (
                  SELECT SKU_NBR, ROUND(ECH_DPTH * ECH_WDTH * ECH_HGHT, 2) AS ITEM_CUBE
                  FROM {SCHN_SKU_ATTR} WHERE LATEST_SKU_CRT_DT_FLG IS TRUE
                )
                SELECT CAST(s.FACTORY_ID AS INT64) AS FACTORY_ID,
                       ROUND(SUM(a.ITEM_CUBE * CAST(
                         CASE WHEN MOD(s.BUY_UNITS, s.BP) = 0 THEN s.BUY_UNITS
                              ELSE CEIL(s.BUY_UNITS / s.BP) * s.BP END AS INT64)), 2) AS FACTORY_CUBE
                FROM src s LEFT JOIN attr a ON s.THD_SKU_NBR = a.SKU_NBR
                WHERE s.FACTORY_ID IS NOT NULL
                GROUP BY s.FACTORY_ID
            """
            cube_df = bq().query(cube_query).to_dataframe()
            factory_cubes = {int(r["FACTORY_ID"]): float(r["FACTORY_CUBE"]) if pd.notna(r["FACTORY_CUBE"]) else 0.0 for _, r in cube_df.iterrows()}
            _upload_cache["factory_cubes"] = factory_cubes

        checks.append({
            "name": "Validation table created",
            "passed": True,
            "detail": table_id.split(".")[-1],
        })

        resp = {"passed": True, "checks": checks, "validation_table": table_id}
        if includes_imports and "FACTORY_ID" in df.columns:
            resp["factory_cubes"] = _upload_cache.get("factory_cubes", {})
        return jsonify(resp)
    except Exception as e:
        logger.exception("Temp table creation error")
        return jsonify({"error": f"Failed to create temp table: {e}"}), 500


@app.route("/api/download_invalid_skus")
def download_invalid_skus():
    invalid_df = _upload_cache.get("invalid_skus")
    if invalid_df is None:
        return jsonify({"error": "No invalid SKU data available"}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = "Invalid SKUs"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    reason_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    cols = list(invalid_df.columns)
    for i, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 14)

    for r_idx, (_, row) in enumerate(invalid_df.iterrows(), 2):
        for c_idx, col_name in enumerate(cols, 1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if col_name == "REASON":
                cell.fill = reason_fill

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="invalid_skus.xlsx",
    )


# ── Section 3: Insert to BigQuery ───────────────────────────────────

def _build_insert_query(source_table: str, container_divisor: int = 2390, includes_imports: bool = True) -> str:
    """Build the enrichment INSERT query using the validated temp table as source."""
    cd = container_divisor
    factory_id_expr = "SAFE_CAST(FACTORY_ID AS INT64) AS FACTORY_ID," if includes_imports else "CAST(NULL AS INT64) AS FACTORY_ID,"
    return f"""
INSERT INTO {EVENTS_SKU_LIST} (
  EVENT_NAME, EVENT_YEAR, THD_SKU_NBR, SISTER_SKU_NBR, BP, BUY_UNITS,
  W1_UNITS, W2_UNITS, W3_UNITS, W4_UNITS, W5_UNITS,
  SUPPLIER, MVNDR_NBR, FACTORY_ID, SKU_NBR, SKU_DESC, IS_SISTER_SKU_FLAG,
  LENGTH, WIDTH, HEIGHT, WEIGHT, ITEM_CUBE,
  DEPT, CLASS, EXT_SUB_CLASS_NBR, SUB_CLASS, SKU_CRT_DT,
  ITEM_CONTAINER, SKU_LEVEL_CONTAINERS,
  FACTORY_CUBE, FACTORY_CONTAINERS, SKU_PCT_OF_FACTORY_CONTAINERS,
  THD_KEY_ID
)
WITH SKU_LIST AS (
  SELECT
    SAFE_CAST(THD_SKU_NBR AS INT64) AS THD_SKU_NBR,
    SAFE_CAST(BUY_UNITS AS INT64) AS BUY_UNITS,
    {factory_id_expr}
    -- Trimmed at ingestion. Whitespace here would otherwise flow all the way through to
    -- the allocation results AND change the THD_KEY_ID fingerprint, so 'Patio ' and
    -- 'Patio' would become two different records.
    NULLIF(TRIM(SAFE_CAST(EVENT_NAME AS STRING)), '') AS EVENT_NAME,
    SAFE_CAST(EVENT_YEAR AS INT64) AS EVENT_YEAR,
    SAFE_CAST(SISTER_SKU_NBR AS INT64) AS SISTER_SKU_NBR,
    SAFE_CAST(BP AS INT64) AS BP,
    SAFE_CAST(WAVE_1 AS INT64) AS W1_UNITS,
    SAFE_CAST(WAVE_2 AS INT64) AS W2_UNITS,
    SAFE_CAST(WAVE_3 AS INT64) AS W3_UNITS,
    SAFE_CAST(WAVE_4 AS INT64) AS W4_UNITS,
    SAFE_CAST(WAVE_5 AS INT64) AS W5_UNITS,
        COUNTIF(
            SAFE_CAST(WAVE_1 AS INT64) IS NOT NULL
            OR SAFE_CAST(WAVE_2 AS INT64) IS NOT NULL
            OR SAFE_CAST(WAVE_3 AS INT64) IS NOT NULL
            OR SAFE_CAST(WAVE_4 AS INT64) IS NOT NULL
            OR SAFE_CAST(WAVE_5 AS INT64) IS NOT NULL
        ) OVER () > 0 AS WAVE_DATA_PRESENT,
    NULLIF(TRIM(SAFE_CAST(SUPPLIER AS STRING)), '') AS SUPPLIER,
    SAFE_CAST(MVNDR_NBR AS INT64) AS MVNDR_NBR,
    NULLIF(TRIM(SAFE_CAST(SKU_DESC AS STRING)), '') AS SKU_DESC
  FROM `{source_table}`
),
SKU_ATTR AS (
  SELECT
    SKU_NBR, SKU_DESC, SKU_CRT_DT,
    ROUND(ECH_DPTH * 12, 2) AS LENGTH,
    ROUND(ECH_WDTH * 12, 2) AS WIDTH,
    ROUND(ECH_HGHT * 12, 2) AS HEIGHT,
    ROUND(ECH_WGHT, 2) AS WEIGHT,
    ROUND((ECH_DPTH * ECH_WDTH * ECH_HGHT), 2) AS ITEM_CUBE,
    DEPT, CLASS, SUB_CLASS, EXT_SUB_CLASS_NBR
  FROM {SCHN_SKU_ATTR}
  WHERE LATEST_SKU_CRT_DT_FLG IS TRUE
),
STAGED_SKU_LIST AS (
  SELECT
    OG.EVENT_NAME, OG.EVENT_YEAR, OG.BP, OG.BUY_UNITS,
    OG.W1_UNITS, OG.W2_UNITS, OG.W3_UNITS, OG.W4_UNITS, OG.W5_UNITS,
    OG.SUPPLIER, OG.MVNDR_NBR, OG.FACTORY_ID,
    OG.THD_SKU_NBR,
    OG.SISTER_SKU_NBR,
    OG.SKU_DESC,
    -- Fallback description, keyed on THD_SKU_NBR. It must come from the THD SKU, never the
    -- proxy: when IS_SISTER_SKU_FLAG is true the proxy is the SISTER SKU, so falling back
    -- to the proxy's description would label the record with the sister's name.
    NULLIF(TRIM(THD.SKU_DESC), '') AS THD_ATTR_SKU_DESC,
    THD.SKU_CRT_DT AS THD_SKU_CRT_DT,
    THD.LENGTH AS THD_LENGTH, THD.WIDTH AS THD_WIDTH,
    THD.HEIGHT AS THD_HEIGHT, THD.WEIGHT AS THD_WEIGHT,
    THD.ITEM_CUBE AS THD_ITEM_CUBE,
    THD.DEPT AS THD_DEPT, THD.CLASS AS THD_CLASS,
    THD.SUB_CLASS AS THD_SUB_CLASS, THD.EXT_SUB_CLASS_NBR AS THD_EXT_SUB_CLASS_NBR,
    SIS.SKU_CRT_DT AS SIS_SKU_CRT_DT,
    DATE_DIFF(CURRENT_DATE(), DATE(THD.SKU_CRT_DT), DAY) AS THD_DAYS,
    DATE_DIFF(CURRENT_DATE(), DATE(SIS.SKU_CRT_DT), DAY) AS SIS_DAYS
  FROM SKU_LIST OG
  LEFT JOIN SKU_ATTR AS THD ON OG.THD_SKU_NBR = THD.SKU_NBR
  LEFT JOIN SKU_ATTR AS SIS ON OG.SISTER_SKU_NBR = SIS.SKU_NBR
),
FINAL_LOGIC_APPLIED AS (
  SELECT *,
    CASE
      WHEN COALESCE(THD_DAYS, 0) < 365 AND COALESCE(SIS_DAYS, 0) >= 365 THEN TRUE
      ELSE FALSE
    END AS IS_SISTER_SKU_FLAG
  FROM STAGED_SKU_LIST
),
ROUNDED AS (
  -- Buy-pack round-up computed ONCE. It used to be repeated inline six times, which is how
  -- the stored BUY_UNITS and the value feeding THD_KEY_ID could drift apart.
  SELECT *,
    SAFE_CAST(
      CASE WHEN MOD(BUY_UNITS, BP) = 0 THEN BUY_UNITS
           ELSE CEIL(BUY_UNITS / BP) * BP
      END AS INT64) AS BUY_UNITS_RND
  FROM FINAL_LOGIC_APPLIED
),
WAVE_SUMMED AS (
  SELECT *,
    IFNULL(W1_UNITS,0) + IFNULL(W2_UNITS,0) + IFNULL(W3_UNITS,0)
      + IFNULL(W4_UNITS,0) + IFNULL(W5_UNITS,0) AS W_SUM
  FROM ROUNDED
),
WAVE_SCALED AS (
  -- Waves must sum to the ROUNDED buy units. BUY_UNITS is rounded UP to a pack multiple, so
  -- uploaded waves (which sum to the pre-round figure) would otherwise leave the wave plan
  -- short by up to BP-1 units and the allocation would ship fewer units by wave than the
  -- record's own BUY_UNITS. This is the source of truth for the wave plan.
  --
  -- Waves are deliberately NOT forced to pack multiples here: the allocation floors to BP
  -- per (DC, wave) downstream, so aligning twice would double-round and lose units.
  SELECT *,
    IF(W_SUM = 0, 0, CAST(FLOOR(IFNULL(W1_UNITS,0) * BUY_UNITS_RND / W_SUM) AS INT64)) AS B1,
    IF(W_SUM = 0, 0, CAST(FLOOR(IFNULL(W2_UNITS,0) * BUY_UNITS_RND / W_SUM) AS INT64)) AS B2,
    IF(W_SUM = 0, 0, CAST(FLOOR(IFNULL(W3_UNITS,0) * BUY_UNITS_RND / W_SUM) AS INT64)) AS B3,
    IF(W_SUM = 0, 0, CAST(FLOOR(IFNULL(W4_UNITS,0) * BUY_UNITS_RND / W_SUM) AS INT64)) AS B4,
    IF(W_SUM = 0, 0, CAST(FLOOR(IFNULL(W5_UNITS,0) * BUY_UNITS_RND / W_SUM) AS INT64)) AS B5,
    -- Remainder lands on the LAST non-empty wave, so rounding never pulls inventory earlier
    -- than planned and never revives a wave the plan left empty. No wave detail at all
    -- (W_SUM = 0) collapses to a single wave 1 only when another uploaded row
    -- contains wave data; a list with no wave data stays wave-free.
    CASE WHEN W_SUM = 0 AND WAVE_DATA_PRESENT THEN 1
         WHEN IFNULL(W5_UNITS,0) > 0 THEN 5
         WHEN IFNULL(W4_UNITS,0) > 0 THEN 4
         WHEN IFNULL(W3_UNITS,0) > 0 THEN 3
         WHEN IFNULL(W2_UNITS,0) > 0 THEN 2
         ELSE 1
    END AS LAST_W
  FROM WAVE_SUMMED
),
WAVE_FINAL AS (
  SELECT * EXCEPT(B1, B2, B3, B4, B5, LAST_W, W_SUM,
                  W1_UNITS, W2_UNITS, W3_UNITS, W4_UNITS, W5_UNITS),
    B1 + IF(LAST_W = 1 AND WAVE_DATA_PRESENT, BUY_UNITS_RND - (B1+B2+B3+B4+B5), 0) AS W1_UNITS,
    B2 + IF(LAST_W = 2, BUY_UNITS_RND - (B1+B2+B3+B4+B5), 0) AS W2_UNITS,
    B3 + IF(LAST_W = 3, BUY_UNITS_RND - (B1+B2+B3+B4+B5), 0) AS W3_UNITS,
    B4 + IF(LAST_W = 4, BUY_UNITS_RND - (B1+B2+B3+B4+B5), 0) AS W4_UNITS,
    B5 + IF(LAST_W = 5, BUY_UNITS_RND - (B1+B2+B3+B4+B5), 0) AS W5_UNITS
  FROM WAVE_SCALED
),
MEASURED AS (
  SELECT *,
    ROUND(SAFE_DIVIDE(THD_ITEM_CUBE, {cd}), 4) AS ITEM_CONTAINER_RAW,
    ROUND(SAFE_DIVIDE(THD_ITEM_CUBE, {cd}) * BUY_UNITS_RND, 4) AS SKU_CONTAINERS_RAW,
    -- Resolved once so the STORED column and the THD_KEY_ID fingerprint use the same string
    COALESCE(SKU_DESC, THD_ATTR_SKU_DESC) AS SKU_DESC_FINAL
  FROM WAVE_FINAL
)
SELECT
  EVENT_NAME, EVENT_YEAR,
  THD_SKU_NBR,
  SISTER_SKU_NBR,
  BP,
  BUY_UNITS_RND AS BUY_UNITS,
  W1_UNITS, W2_UNITS, W3_UNITS, W4_UNITS, W5_UNITS,
  SUPPLIER, MVNDR_NBR, FACTORY_ID,
  CASE WHEN IS_SISTER_SKU_FLAG = FALSE THEN THD_SKU_NBR
       ELSE SISTER_SKU_NBR
  END AS SKU_NBR,
  SKU_DESC_FINAL AS SKU_DESC,
  IS_SISTER_SKU_FLAG,
  THD_LENGTH AS LENGTH, THD_WIDTH AS WIDTH,
  THD_HEIGHT AS HEIGHT, THD_WEIGHT AS WEIGHT,
  THD_ITEM_CUBE AS ITEM_CUBE,
  THD_DEPT AS DEPT, THD_CLASS AS CLASS,
  THD_EXT_SUB_CLASS_NBR AS EXT_SUB_CLASS_NBR,
  THD_SUB_CLASS AS SUB_CLASS,
  THD_SKU_CRT_DT AS SKU_CRT_DT,
  IF(FACTORY_ID IS NOT NULL, ITEM_CONTAINER_RAW, NULL) AS ITEM_CONTAINER,
  IF(FACTORY_ID IS NOT NULL, SKU_CONTAINERS_RAW, NULL) AS SKU_LEVEL_CONTAINERS,
  IF(FACTORY_ID IS NOT NULL,
     ROUND(SUM(THD_ITEM_CUBE * BUY_UNITS_RND) OVER (PARTITION BY FACTORY_ID), 2),
     NULL) AS FACTORY_CUBE,
  IF(FACTORY_ID IS NOT NULL,
     ROUND(SUM(SKU_CONTAINERS_RAW) OVER (PARTITION BY FACTORY_ID), 2),
     NULL) AS FACTORY_CONTAINERS,
  IF(FACTORY_ID IS NOT NULL,
     SAFE_DIVIDE(SKU_CONTAINERS_RAW,
                 SUM(SKU_CONTAINERS_RAW) OVER (PARTITION BY FACTORY_ID)) * 100,
     NULL) AS SKU_PCT_OF_FACTORY_CONTAINERS,
  -- THE RECORD GRAIN. Every field aliased to its COLUMN name on purpose: TO_JSON_STRING
  -- keys the JSON by field name, so an unaliased or renamed expression hashes differently
  -- than a recompute from the stored row. Must stay byte-identical to the standalone
  -- enrichment query in Allocation Automation/events_sku_list_upload_enrichment.sql.
  FARM_FINGERPRINT(TO_JSON_STRING(STRUCT(
    EVENT_NAME AS EVENT_NAME,
    EVENT_YEAR AS EVENT_YEAR,
    THD_SKU_NBR AS THD_SKU_NBR,
    SISTER_SKU_NBR AS SISTER_SKU_NBR,
    SKU_DESC_FINAL AS SKU_DESC,
    SUPPLIER AS SUPPLIER,
    MVNDR_NBR AS MVNDR_NBR,
    FACTORY_ID AS FACTORY_ID,
    BP AS BP,
    BUY_UNITS_RND AS BUY_UNITS,
    W1_UNITS AS W1_UNITS, W2_UNITS AS W2_UNITS, W3_UNITS AS W3_UNITS,
    W4_UNITS AS W4_UNITS, W5_UNITS AS W5_UNITS
  ))) AS THD_KEY_ID
FROM MEASURED
"""


@app.route("/api/check_insert_status", methods=["POST"])
def check_insert_status():
    """Check if current upload already exists in EVENTS_SKU_LIST."""
    body = request.get_json(silent=True) or {}
    df = _upload_cache.get("df")
    upload_count = len(df) if df is not None and len(df) > 0 else None

    event_name = _upload_cache.get("event_name") or body.get("event_name", "")
    event_year = _upload_cache.get("event_year") or body.get("event_year")
    includes_imports = _upload_cache.get("includes_imports", body.get("includes_imports", False))
    if not event_name or event_year is None:
        return jsonify({"already_inserted": False})
    try:
        event_year = int(event_year)
    except (TypeError, ValueError):
        return jsonify({"already_inserted": False})

    try:
        query = f"""
            SELECT COUNT(*) AS cnt,
                   COUNTIF(FACTORY_ID IS NOT NULL AND FACTORY_ID != 0) AS factory_cnt
            FROM {EVENTS_SKU_LIST}
            WHERE UPPER(EVENT_NAME) = @event_name AND EVENT_YEAR = @event_year
        """
        job_config = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("event_name", "STRING", event_name.upper()),
            bigquery.ScalarQueryParameter("event_year", "INT64", event_year),
        ])
        row = next(iter(bq().query(query, job_config=job_config).result()))
        cnt = row.cnt
        existing_is_import = row.factory_cnt > 0

        if cnt == 0:
            return jsonify({"already_inserted": False})

        # Fetch factory containers from existing data
        factory_cubes = {}
        if includes_imports:
            try:
                fc_query = f"""
                    SELECT FACTORY_ID, MAX(FACTORY_CUBE) AS FACTORY_CUBE
                    FROM {EVENTS_SKU_LIST}
                    WHERE UPPER(EVENT_NAME) = @event_name AND EVENT_YEAR = @event_year
                      AND FACTORY_ID IS NOT NULL
                    GROUP BY FACTORY_ID
                """
                fc_rows = bq().query(fc_query, job_config=job_config).result()
                factory_cubes = {int(r.FACTORY_ID): float(r.FACTORY_CUBE or 0) for r in fc_rows}
                _upload_cache["factory_cubes"] = factory_cubes
            except Exception:
                logger.exception("Factory cube fetch from EVENTS_SKU_LIST failed")

        upload_count = len(df)
        # Different type (import vs domestic) — allow replace
        type_differs = existing_is_import != includes_imports
        if type_differs:
            existing_type = "import" if existing_is_import else "domestic"
            upload_type = "import" if includes_imports else "domestic"
            return jsonify({
                "already_inserted": False,
                "exists_different": True,
                "message": f"'{event_name}' ({event_year}) has {cnt:,} {existing_type} rows but your upload is {upload_type} ({upload_count:,} rows). Would you like to replace?",
                "factory_cubes": factory_cubes,
            })

        # No upload in cache — just check if data exists
        if upload_count is None:
            return jsonify({
                "already_inserted": True,
                "message": f"'{event_name}' ({event_year}) already inserted — {cnt:,} rows in BigQuery.",
                "factory_cubes": factory_cubes,
            })

        # Same type, same row count — already inserted
        if cnt == upload_count:
            return jsonify({
                "already_inserted": True,
                "message": f"'{event_name}' ({event_year}) already inserted — {cnt:,} rows match your upload.",
                "factory_cubes": factory_cubes,
            })

        # Same type, different row count
        return jsonify({
            "already_inserted": False,
            "exists_different": True,
            "message": f"'{event_name}' ({event_year}) already has {cnt:,} rows but your upload has {upload_count:,}. Would you like to replace?",
            "factory_cubes": factory_cubes,
        })
    except Exception as e:
        logger.exception("Insert status check failed")
        return jsonify({"already_inserted": False})


@app.route("/api/insert", methods=["POST"])
def insert_to_bq():
    df = _upload_cache.get("df")
    if df is None or len(df) == 0:
        return jsonify({"error": "No validated data. Please upload first."}), 400

    event_name = _upload_cache.get("event_name", "")
    event_year = _upload_cache.get("event_year", 2026)

    body = request.get_json(silent=True) or {}
    overwrite = body.get("overwrite", False)

    # Check event doesn't already exist in EVENTS_SKU_LIST
    try:
        dup_query = f"""
            SELECT COUNT(*) AS cnt FROM {EVENTS_SKU_LIST}
            WHERE UPPER(EVENT_NAME) = @event_name AND EVENT_YEAR = @event_year
        """
        job_config = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("event_name", "STRING", event_name.upper()),
            bigquery.ScalarQueryParameter("event_year", "INT64", event_year),
        ])
        cnt = next(iter(bq().query(dup_query, job_config=job_config).result())).cnt
        if cnt > 0 and not overwrite:
            return jsonify({
                "success": False,
                "exists": True,
                "message": f"'{event_name}' ({event_year}) already has {cnt:,} rows in EVENTS_SKU_LIST.",
                "row_count": cnt,
            })
        if cnt > 0 and overwrite:
            delete_query = f"""
                DELETE FROM {EVENTS_SKU_LIST}
                WHERE UPPER(EVENT_NAME) = @event_name AND EVENT_YEAR = @event_year
            """
            bq().query(delete_query, job_config=job_config).result()
            logger.info(f"Deleted {cnt} existing rows for {event_name} {event_year}")
    except Exception as e:
        logger.exception("Event duplicate check failed")
        return jsonify({"error": f"Failed to check for existing event: {e}"}), 500

    validation_table = _upload_cache.get("validation_table")

    # Upload df to temp table if not already staged
    if not validation_table:
        safe_event = _sanitize_table_name(event_name) if event_name else "UPLOAD"
        table_id = f"{PROJECT_ID}.{TEMP_DATASET}.VALIDATION_{safe_event}_{event_year}"
        try:
            jc = bigquery.LoadJobConfig(
                write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
            )
            bq().load_table_from_dataframe(df, table_id, job_config=jc).result()
            validation_table = table_id
            _upload_cache["validation_table"] = table_id
        except Exception as e:
            logger.exception("Failed to create temp table")
            return jsonify({"error": f"Failed to stage data: {e}"}), 500

    container_divisor = int(body.get("container_divisor", CONTAINER_DIVISOR))
    if container_divisor <= 0:
        container_divisor = CONTAINER_DIVISOR

    includes_imports = _upload_cache.get("includes_imports", False)
    row_count = len(df)

    try:
        insert_sql = _build_insert_query(validation_table, container_divisor, includes_imports)
        job = bq().query(insert_sql)
        job.result()
        rows_inserted = job.num_dml_affected_rows or row_count

        # Keep the validated frame available for vendor matching and SKU-level
        # DC overrides after the event has been inserted.
        _upload_cache.pop("validation_table", None)
        return jsonify({
            "success": True,
            "message": f"Inserted {rows_inserted} enriched rows into EVENTS_SKU_LIST.",
            "row_count": rows_inserted,
        })
    except Exception as e:
        logger.exception("Insert error")
        return jsonify({"error": str(e)}), 500


# ── Section 4: Vendor Strategy Table ────────────────────────────────

@app.route("/api/vendor_strategy")
def api_vendor_strategy():
    try:
        query = f"SELECT VENDOR, ASMT_ID, DC_COUNT, DC_LIST, DC_NM_LIST FROM {VENDOR_STRATEGY} ORDER BY VENDOR"
        rows = [dict(r) for r in bq().query(query).result()]
        return jsonify({"data": rows})
    except Exception as e:
        logger.exception("Vendor strategy error")
        return jsonify({"error": str(e)}), 500


@app.route("/api/match_vendor_strategy", methods=["POST"])
def api_match_vendor_strategy():
    """Match uploaded suppliers to VENDOR_ALIGNED_STRATEGY using fuzzy LIKE."""
    body = request.get_json(silent=True) or {}
    event_name = body.get("event_name") or _upload_cache.get("event_name", "")

    # Try upload cache first, fall back to BQ
    df = _upload_cache.get("df")
    sku_counts = {}
    if df is not None and "SUPPLIER" in df.columns:
        suppliers = df["SUPPLIER"].dropna().unique().tolist()
        # Count on THD_KEY, not THD_SKU_NBR: the same SKU legitimately appears
        # on more than one row (different factory, buy pack, or wave), and
        # counting the SKU number alone silently collapses those into one.
        key_cols = [c for c in _determine_thd_key(df, _upload_cache.get("includes_imports", False))
                    if c in df.columns]
        work = df.dropna(subset=["SUPPLIER"]).copy()
        work["_THD_KEY"] = work[key_cols].astype(str).agg("|".join, axis=1)
        sku_counts = work.groupby("SUPPLIER")["_THD_KEY"].nunique().to_dict()
    elif event_name:
        try:
            q = f"""SELECT SUPPLIER, COUNT(DISTINCT THD_SKU_NBR) AS SKU_COUNT
                    FROM {EVENTS_SKU_LIST} WHERE EVENT_NAME = @ev AND SUPPLIER IS NOT NULL
                    GROUP BY SUPPLIER"""
            jc = bigquery.QueryJobConfig(query_parameters=[
                bigquery.ScalarQueryParameter("ev", "STRING", event_name),
            ])
            rows = list(bq().query(q, job_config=jc).result())
            suppliers = [r.SUPPLIER for r in rows]
            sku_counts = {r.SUPPLIER: r.SKU_COUNT for r in rows}
        except Exception as e:
            return jsonify({"error": f"Failed to read suppliers from BQ: {e}"}), 500
    else:
        return jsonify({"error": "No uploaded data available. Please upload a file first."}), 400

    if not suppliers:
        return jsonify({"error": "No suppliers found."}), 400

    try:
        vs_query = f"SELECT VENDOR, ASMT_ID, DC_COUNT, DC_LIST, DC_NM_LIST FROM {VENDOR_STRATEGY}"
        vs_rows = [dict(r) for r in bq().query(vs_query).result()]

        matches = []
        unmatched = []
        for supplier in sorted(suppliers):
            sup_upper = supplier.upper().strip()
            sku_count = int(sku_counts.get(supplier, 0))
            matched = False
            for v in vs_rows:
                vendor_upper = v["VENDOR"].upper().strip()
                if vendor_upper in sup_upper:
                    matches.append({
                        "SUPPLIER": supplier,
                        "VENDOR": v["VENDOR"],
                        "ASMT_ID": v["ASMT_ID"],
                        "DC_COUNT": v["DC_COUNT"],
                        "DC_LIST": v["DC_LIST"],
                        "DC_NM_LIST": v["DC_NM_LIST"],
                        "SKU_COUNT": sku_count,
                    })
                    matched = True
                    break
            if not matched:
                other = next((v for v in vs_rows if v["VENDOR"].upper() == "OTHER"), None)
                if other:
                    matches.append({
                        "SUPPLIER": supplier,
                        "VENDOR": "OTHER",
                        "ASMT_ID": other["ASMT_ID"],
                        "DC_COUNT": other["DC_COUNT"],
                        "DC_LIST": other["DC_LIST"],
                        "DC_NM_LIST": other["DC_NM_LIST"],
                        "SKU_COUNT": sku_count,
                    })
                else:
                    unmatched.append(supplier)

        unique_strategies = {m["ASMT_ID"] for m in matches}

        return jsonify({
            "matches": matches,
            "unmatched": unmatched,
            "strategy_count": len(unique_strategies),
            "supplier_count": len(suppliers),
            "sku_count": sum(m["SKU_COUNT"] for m in matches),
        })
    except Exception as e:
        logger.exception("Match vendor strategy error")
        return jsonify({"error": str(e)}), 500


# ── Section 4b: DFC Cost Model Submission ────────────────────────────

@app.route("/api/vendor_skus")
def api_vendor_skus():
    """Return one supplier's validated SKU rows for the expandable drill-down."""
    df = _upload_cache.get("df")
    supplier = request.args.get("supplier", "").strip()
    if df is None or not supplier:
        return jsonify({"error": "A validated upload and supplier are required."}), 400
    if "SUPPLIER" not in df.columns:
        return jsonify({"error": "The validated upload has no SUPPLIER column."}), 400

    try:
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("page_size", 50)), 1), 100)
    except ValueError:
        return jsonify({"error": "page and page_size must be numbers."}), 400

    supplier_mask = df["SUPPLIER"].fillna("").astype(str).str.strip().str.upper() == supplier.upper()
    rows_df = df.loc[supplier_mask].copy()
    key_cols = [c for c in _determine_thd_key(df, _upload_cache.get("includes_imports", False)) if c in df.columns]
    rows_df["THD_KEY"] = rows_df[key_cols].fillna("__NULL__").astype(str).agg("|".join, axis=1)
    total = len(rows_df)
    start = (page - 1) * page_size

    def numeric_value(value, default=0):
        try:
            return float(str(value).replace(",", "").strip())
        except (TypeError, ValueError):
            return default

    item_cubes = {}
    sku_values = [int(numeric_value(value)) for value in rows_df["THD_SKU_NBR"].dropna().tolist()]
    if sku_values:
        cube_query = f"""
            SELECT THD_SKU_NBR AS SKU_NBR, MAX(ITEM_CUBE) AS ITEM_CUBE
            FROM {EVENTS_SKU_LIST}
            WHERE UPPER(EVENT_NAME) = @event_name
              AND EVENT_YEAR = @event_year
              AND THD_SKU_NBR IN UNNEST(@sku_list)
            GROUP BY THD_SKU_NBR
        """
        cube_config = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("event_name", "STRING", _upload_cache.get("event_name", "").upper()),
            bigquery.ScalarQueryParameter("event_year", "INT64", int(_upload_cache.get("event_year", 2026))),
            bigquery.ArrayQueryParameter("sku_list", "INT64", sorted(set(sku_values))),
        ])
        for cube_row in bq().query(cube_query, job_config=cube_config).result():
            item_cubes[int(cube_row.SKU_NBR)] = float(cube_row.ITEM_CUBE or 0)

        missing_cubes = sorted(set(sku_values) - set(item_cubes))
        if missing_cubes:
            fallback_query = f"""
                SELECT SKU_NBR, ROUND(ECH_DPTH * ECH_WDTH * ECH_HGHT, 2) AS ITEM_CUBE
                FROM {SCHN_SKU_ATTR}
                WHERE SKU_NBR IN UNNEST(@sku_list) AND LATEST_SKU_CRT_DT_FLG IS TRUE
            """
            fallback_config = bigquery.QueryJobConfig(query_parameters=[
                bigquery.ArrayQueryParameter("sku_list", "INT64", missing_cubes),
            ])
            for cube_row in bq().query(fallback_query, job_config=fallback_config).result():
                item_cubes[int(cube_row.SKU_NBR)] = float(cube_row.ITEM_CUBE or 0)

    rows = []
    for _, row in rows_df.iloc[start:start + page_size].iterrows():
        item = {}
        for col in ["THD_KEY", "THD_SKU_NBR", "SISTER_SKU_NBR", "SKU_DESC", "BP", "BUY_UNITS", "WAVE_1", "WAVE_2", "WAVE_3", "WAVE_4", "WAVE_5"]:
            if col not in row.index:
                continue
            value = row[col]
            item[col] = "" if pd.isna(value) else str(value)
        sku_nbr = int(numeric_value(row["THD_SKU_NBR"])) if pd.notna(row.get("THD_SKU_NBR")) else None
        buy_units = numeric_value(row.get("BUY_UNITS")) if pd.notna(row.get("BUY_UNITS")) else 0
        item["TOTAL_UNITS"] = str(int(buy_units)) if buy_units else ""
        item["TOTAL_CUBE"] = f"{item_cubes.get(sku_nbr, 0) * buy_units:.2f}" if sku_nbr else ""
        rows.append(item)
    return jsonify({"rows": rows, "page": page, "page_size": page_size, "total": total, "pages": (total + page_size - 1) // page_size})


@app.route("/api/cost_model_preview", methods=["POST"])
def api_cost_model_preview():
    """Return SKU breakdown by sister flag for charts + table preview."""
    body = request.get_json(silent=True) or {}
    event_name = body.get("event_name", "")
    if not event_name:
        return jsonify({"error": "event_name is required"}), 400

    try:
        q = f"""
            SELECT
              CAST(SKU_NBR AS STRING) AS sku_nbr,
              SUM(BUY_UNITS) AS buy_qty,
              MAX(CASE WHEN IS_SISTER_SKU_FLAG THEN TRUE ELSE FALSE END) AS IS_SISTER_SKU_FLAG
            FROM {EVENTS_SKU_LIST}
            WHERE EVENT_NAME = @ev
            GROUP BY SKU_NBR
            ORDER BY buy_qty DESC
        """
        jc = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("ev", "STRING", event_name),
        ])
        rows = [dict(r) for r in bq().query(q, job_config=jc).result()]

        thd_units = sum(r["buy_qty"] for r in rows if not r["IS_SISTER_SKU_FLAG"])
        sister_units = sum(r["buy_qty"] for r in rows if r["IS_SISTER_SKU_FLAG"])
        thd_count = sum(1 for r in rows if not r["IS_SISTER_SKU_FLAG"])
        sister_count = sum(1 for r in rows if r["IS_SISTER_SKU_FLAG"])

        thd_skus = {r["sku_nbr"] for r in rows if not r["IS_SISTER_SKU_FLAG"]}
        sister_skus = {r["sku_nbr"] for r in rows if r["IS_SISTER_SKU_FLAG"]}
        overlap_skus = sorted(thd_skus & sister_skus)

        return jsonify({
            "rows": rows,
            "thd_units": thd_units,
            "sister_units": sister_units,
            "thd_count": thd_count,
            "sister_count": sister_count,
            "overlap_skus": overlap_skus,
        })
    except Exception as e:
        logger.exception("cost_model_preview error")
        return jsonify({"error": str(e)}), 500


@app.route("/api/submit_cost_model", methods=["POST"])
def api_submit_cost_model():
    """Insert per-SKU rows into DFC_COST_MODEL_SUBMISSION from EVENTS_SKU_LIST."""
    body = request.get_json(silent=True) or {}
    event_name = body.get("event_name", "")
    if not event_name:
        return jsonify({"error": "event_name is required"}), 400

    try:
        # Derive event_year from EVENTS_SKU_LIST
        year_q = f"SELECT DISTINCT EVENT_YEAR FROM {EVENTS_SKU_LIST} WHERE EVENT_NAME = @ev LIMIT 1"
        jc = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("ev", "STRING", event_name),
        ])
        year_rows = list(bq().query(year_q, job_config=jc).result())
        if not year_rows:
            return jsonify({"error": f"No rows found for event '{event_name}'"}), 400
        event_year = year_rows[0].EVENT_YEAR

        project_name = f"{event_year} {event_name}"
        key = f"{CURRENT_USER}-{project_name}-{project_name}"

        # Check for duplicate submission
        dup_q = f"SELECT COUNT(*) AS cnt FROM {DFC_COST_MODEL_SUBMISSION} WHERE `key` = @key"
        dup_jc = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("key", "STRING", key),
        ])
        dup_rows = list(bq().query(dup_q, job_config=dup_jc).result())
        if dup_rows and dup_rows[0].cnt > 0:
            return jsonify({"error": f"This event has already been submitted ({dup_rows[0].cnt} rows exist). Delete the previous submission before resubmitting."}), 409

        # Look up email from ASSOC table
        email_q = "SELECT EMAIL_ADDR_TXT FROM `pr-edw-views-thd.ASSOC.ASSOC` WHERE USER_ID = @ldap LIMIT 1"
        email_jc = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("ldap", "STRING", CURRENT_USER),
        ])
        email_rows = list(bq().query(email_q, job_config=email_jc).result())
        email = email_rows[0].EMAIL_ADDR_TXT if email_rows else f"{CURRENT_USER}@homedepot.com"

        submission_sql = f"""
            INSERT INTO {DFC_COST_MODEL_SUBMISSION}
              (date_added, `key`, project_name, bucket, sku_nbr, buy_qty,
               email, target_dc_count, dc_inclusions, dc_exclusions)
            SELECT
              CURRENT_DATE() AS date_added,
              @key AS `key`,
              @project_name AS project_name,
              @project_name AS bucket,
              SAFE_CAST(SKU_NBR AS STRING) AS sku_nbr,
              CAST(SUM(BUY_UNITS) AS STRING) AS buy_qty,
              @email AS email,
              CAST(NULL AS STRING) AS target_dc_count,
              CAST(NULL AS STRING) AS dc_inclusions,
              CAST(NULL AS STRING) AS dc_exclusions
            FROM {EVENTS_SKU_LIST}
            WHERE EVENT_NAME = @event_name AND EVENT_YEAR = @event_year
            GROUP BY SKU_NBR
        """
        sub_config = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("key", "STRING", key),
            bigquery.ScalarQueryParameter("project_name", "STRING", project_name),
            bigquery.ScalarQueryParameter("email", "STRING", email),
            bigquery.ScalarQueryParameter("event_name", "STRING", event_name),
            bigquery.ScalarQueryParameter("event_year", "INT64", event_year),
        ])
        job = bq().query(submission_sql, job_config=sub_config)
        job.result()
        rows_inserted = job.num_dml_affected_rows or 0
        logger.info(f"Inserted {rows_inserted} rows into DFC_COST_MODEL_SUBMISSION")
        return jsonify({"success": True, "message": f"Submitted {rows_inserted} SKUs to DFC Cost Model.", "row_count": rows_inserted})
    except Exception as e:
        logger.exception("submit_cost_model error")
        return jsonify({"error": str(e)}), 500


@app.route("/api/delete_cost_model", methods=["POST"])
def api_delete_cost_model():
    body = request.get_json(silent=True) or {}
    event_name = body.get("event_name", "")
    if not event_name:
        return jsonify({"error": "event_name is required"}), 400

    try:
        year_q = f"SELECT DISTINCT EVENT_YEAR FROM {EVENTS_SKU_LIST} WHERE EVENT_NAME = @ev LIMIT 1"
        jc = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("ev", "STRING", event_name),
        ])
        year_rows = list(bq().query(year_q, job_config=jc).result())
        if not year_rows:
            return jsonify({"error": f"No rows found for event '{event_name}'"}), 400
        event_year = year_rows[0].EVENT_YEAR

        project_name = f"{event_year} {event_name}"
        key = f"{CURRENT_USER}-{project_name}-{project_name}"

        del_sql = f"DELETE FROM {DFC_COST_MODEL_SUBMISSION} WHERE `key` = @key"
        del_config = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("key", "STRING", key),
        ])
        job = bq().query(del_sql, job_config=del_config)
        job.result()
        rows_deleted = job.num_dml_affected_rows or 0
        logger.info(f"Deleted {rows_deleted} rows from DFC_COST_MODEL_SUBMISSION for key={key}")
        return jsonify({"success": True, "message": f"Deleted {rows_deleted} previous row(s)."})
    except Exception as e:
        logger.exception("delete_cost_model error")
        return jsonify({"error": str(e)}), 500


def _resolve_is_import(event_name: str, default=None):
    """Authoritative import/domestic check. Prefers this session's own upload cache
    (fast, no BigQuery round-trip); otherwise falls back to what's already
    persisted in EVENTS_SKU_LIST for this event — true regardless of whether THIS
    session did the upload/insert (e.g. reusing an event someone already loaded).
    Returns `default` (typically None) when neither source has an answer."""
    if "includes_imports" in _upload_cache:
        return _upload_cache["includes_imports"]
    if not event_name:
        return default
    try:
        query = f"""
            SELECT COUNTIF(FACTORY_ID IS NOT NULL AND FACTORY_ID != 0) > 0 AS is_import
            FROM {EVENTS_SKU_LIST}
            WHERE UPPER(EVENT_NAME) = @event_name
        """
        job_config = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("event_name", "STRING", event_name.upper()),
        ])
        rows = list(bq().query(query, job_config=job_config).result())
        if not rows or rows[0].is_import is None:
            return default
        return bool(rows[0].is_import)
    except Exception:
        logger.exception(f"Failed to resolve is_import from EVENTS_SKU_LIST for event_name={event_name}")
        return default


def _resolve_sku_grp(run_id: str, event_name: str, default=None):
    """The catalog/assortment tool builds its own SKU_GRP key from the event name
    (e.g. typed as "3. Example Import DC-Count", published as
    "...-3 EXAMPLE IMPORT DCCOUNT-..."), stripping punctuation along the way — so
    the caller's own event_name string is never assumed to BE the SKU_GRP. A plain
    `UPPER(SKU_GRP) LIKE '%event_name%'` (the pattern used elsewhere in this
    codebase, e.g. run_multi_dc_allocation.sql's resolved_sku_grp) breaks on exactly
    this kind of punctuation difference, so both sides are stripped to bare
    alphanumerics before matching. Returns `default` if nothing matches."""
    if not run_id or not event_name:
        return default
    try:
        query = f"""
            SELECT SKU_GRP
            FROM {CATALOG_RUN_ANALYTICS}
            WHERE RUN_ID = @run_id
              AND REGEXP_REPLACE(UPPER(SKU_GRP), r'[^A-Z0-9]', '')
                  LIKE CONCAT('%', REGEXP_REPLACE(UPPER(@event_name), r'[^A-Z0-9]', ''), '%')
            LIMIT 1
        """
        job_config = bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("run_id", "STRING", run_id),
            bigquery.ScalarQueryParameter("event_name", "STRING", event_name),
        ])
        rows = list(bq().query(query, job_config=job_config).result())
        if not rows:
            return default
        return rows[0].SKU_GRP
    except Exception:
        logger.exception(f"Failed to resolve sku_grp for run_id={run_id} event_name={event_name}")
        return default


# ── Section 5: Assortment ID Determination ──────────────────────────

def _normalize_strategy(body: dict) -> str:
    strategy = body.get("strategy", "")
    if strategy == "DC_SELECTION":
        dc_counts = body.get("dc_counts", [])
        strategy = "SINGLE_DC" if len(dc_counts) == 1 else "MULTI_DC"
        body["strategy"] = strategy
    return strategy


@app.route("/api/determine_assortment", methods=["POST"])
def api_determine_assortment():
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "Invalid JSON body"}), 400

    strategy = _normalize_strategy(body)
    if strategy not in STRATEGY_KEYS:
        return jsonify({"error": f"Invalid strategy: {strategy}"}), 400

    # Import/domestic is decided by what was actually uploaded/persisted, not by
    # whatever the client happens to send — the client-side toggle is only set once
    # (at template download) and can go stale (e.g. after a page refresh, or when
    # reusing an event that already exists in EVENTS_SKU_LIST without re-uploading).
    if "is_import" in body:
        body["is_import"] = _resolve_is_import(body.get("event_name", ""), default=body["is_import"])

    # SKU_GRP is resolved from the catalog run itself, never trusted as typed —
    # the catalog tool builds its own key from the event name and strips
    # punctuation along the way (e.g. "3. Example Import DC-Count" becomes
    # "...3 EXAMPLE IMPORT DCCOUNT..."), so a manually-entered value is unreliable.
    if "sku_grp" in body:
        body["sku_grp"] = _resolve_sku_grp(
            body.get("run_id", ""), body.get("event_name", ""), default=body["sku_grp"]
        )

    result = determine_assortment_ids(bq(), strategy, body)
    if result.get("error"):
        return jsonify(result), 500
    return jsonify(result)


@app.route("/api/determine_assortment_start", methods=["POST"])
def api_determine_assortment_start():
    """Async counterpart to /api/determine_assortment for MULTI_DC — submits the
    procedure call without waiting on it, since the dynamic sweep can run from
    ~40 seconds to several minutes, longer than a single HTTP request reliably
    survives. Returns a job_id to poll via /api/determine_assortment_status
    instead. Other strategies are fast enough to just run synchronously here,
    reported through the same {job_id: null, sync_result: {...}} shape so the
    frontend has one contract regardless of strategy.
    """
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "Invalid JSON body"}), 400

    strategy = _normalize_strategy(body)
    if strategy not in STRATEGY_KEYS:
        return jsonify({"error": f"Invalid strategy: {strategy}"}), 400

    if "is_import" in body:
        body["is_import"] = _resolve_is_import(body.get("event_name", ""), default=body["is_import"])
    if "sku_grp" in body:
        body["sku_grp"] = _resolve_sku_grp(
            body.get("run_id", ""), body.get("event_name", ""), default=body["sku_grp"]
        )

    if strategy == "MULTI_DC":
        started = start_multi_dc(bq(), body)
        if started.get("error"):
            return jsonify(started), 500
        if started["job_id"] is None:
            # No run_id supplied — nothing submitted, existing results already
            # sit in the output tables, so just fetch them now.
            payload = fetch_multi_dc_results(bq())
            return jsonify({"job_id": None, "sync_result": payload, "error": None})
        return jsonify(started)

    result = determine_assortment_ids(bq(), strategy, body)
    if result.get("error"):
        return jsonify(result), 500
    return jsonify({"job_id": None, "sync_result": result, "error": None})


@app.route("/api/determine_assortment_status", methods=["GET"])
def api_determine_assortment_status():
    """Poll a job_id returned by /api/determine_assortment_start."""
    job_id = request.args.get("job_id", "")
    if not job_id:
        return jsonify({"error": "job_id is required"}), 400

    client = bq()
    try:
        job = client.get_job(job_id, location="US")
    except Exception as e:
        logger.warning(f"Could not look up job {job_id}: {e}")
        return jsonify({"done": False})

    if job.state != "DONE":
        return jsonify({"done": False})
    if job.error_result:
        message = job.error_result.get("message", "Assortment determination failed")
        return jsonify({"done": True, "error": message})

    payload = fetch_multi_dc_results(client)
    return jsonify({"done": True, "sync_result": payload, "error": None})


# ── Section 6: Run Allocation ───────────────────────────────────────

@app.route("/api/available_dc_counts", methods=["POST"])
def api_available_dc_counts():
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "Invalid JSON body"}), 400
    run_id = body.get("run_id", "")
    sku_grp = body.get("sku_grp", "")
    if not run_id or not sku_grp:
        return jsonify({"error": "run_id and sku_grp are required"}), 400
    result = fetch_available_dc_counts(bq(), run_id, sku_grp)
    if not result.get("success"):
        return jsonify(result), 500
    return jsonify(result)


@app.route("/api/run_allocation", methods=["POST"])
def api_run_allocation():
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "Invalid JSON body"}), 400

    _normalize_strategy(body)

    if "is_import" in body:
        body["is_import"] = _resolve_is_import(body.get("event_name", ""), default=body["is_import"])

    if "sku_grp" in body:
        body["sku_grp"] = _resolve_sku_grp(
            body.get("run_id", ""), body.get("event_name", ""), default=body["sku_grp"]
        )

    result = run_allocation(bq(), body)
    if not result.get("success"):
        return jsonify(result), 500
    return jsonify(result)


# ── Section 7: Results ──────────────────────────────────────────────

@app.route("/api/results")
def api_results():
    try:
        page = int(request.args.get("page", 1))
        page_size = min(int(request.args.get("page_size", 50)), 200)
        sort = request.args.get("sort", "SKU_NBR")
        direction = request.args.get("dir", "ASC").upper()
        data = fetch_results(bq(), page, page_size, sort, direction)
        return jsonify(data)
    except Exception as e:
        logger.exception("Results error")
        return jsonify({"error": str(e)}), 500


@app.route("/api/results_summary")
def api_results_summary():
    try:
        summary = fetch_summary(bq())
        return jsonify(summary)
    except Exception as e:
        logger.exception("Summary error")
        return jsonify({"error": str(e)}), 500


@app.route("/api/results_validation")
def api_results_validation():
    try:
        checks = validate_results(bq())
        return jsonify({"checks": checks})
    except Exception as e:
        logger.exception("Validation error")
        return jsonify({"error": str(e)}), 500


@app.route("/api/factory_summary")
def api_factory_summary():
    try:
        divisor = float(request.args.get("divisor", 2390))
        data = fetch_factory_summary(bq(), divisor)
        return jsonify({"data": data})
    except Exception as e:
        logger.exception("Factory summary error")
        return jsonify({"error": str(e)}), 500


@app.route("/api/export_results")
def api_export_results():
    try:
        query = f"""
            SELECT THD_SKU_NBR, SISTER_SKU_NBR, SKU_DESC,
                   SUPPLIER, MVNDR_NBR, FACTORY_ID,
                   BP, BUY_UNITS,
                   OG_W1_UNITS, OG_W2_UNITS, OG_W3_UNITS, OG_W4_UNITS, OG_W5_UNITS,
                   SKU_NBR, DC_NBR,
                   DFC_PCT, DFC_UNITS, DFC_W1_UNITS, DFC_W2_UNITS,
                   DFC_W3_UNITS, DFC_W4_UNITS, DFC_W5_UNITS,
                   ITEM_CUBE, RACK_TYPE, FACTORY_CUBE, FACTORY_CONTAINERS
            FROM {FINAL_ALLOCATIONS}
            ORDER BY SKU_NBR, DC_NBR
        """
        df = bq().query(query).to_dataframe()
        buf = io.BytesIO()
        df.to_csv(buf, index=False)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"allocation_results_{datetime.date.today().isoformat()}.csv",
        )
    except Exception as e:
        logger.exception("Export error")
        return jsonify({"error": str(e)}), 500


# ── Startup ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    bq()  # warm up credentials on startup
    app.run(host="127.0.0.1", port=8080, debug=False)
